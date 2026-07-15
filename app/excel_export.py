"""Exports the current database back into the same 4-sheet layout
excel_import.py reads (tram / nhan vien / muc / thiet bi), so the round trip
is: Xuất Excel -> edit in Excel (add/remove/rename rows freely) -> Nhập từ
Excel to bring the changes back in.
"""
import io

from openpyxl import Workbook
from openpyxl.styles import Font
from sqlalchemy.orm import Session, joinedload

from . import models

HEADER_FONT = Font(bold=True)


def _write_sheet(wb: Workbook, title: str, headers: list[str], rows: list[list], col_widths: list[int] | None = None):
    ws = wb.create_sheet(title)
    ws.append(headers)
    for cell in ws[1]:
        cell.font = HEADER_FONT
    ws.freeze_panes = "A2"
    for row in rows:
        ws.append(row)
    widths = col_widths or [18] * len(headers)
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
    return ws


def build_workbook(db: Session) -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)  # drop the default blank sheet

    stations = db.query(models.Station).order_by(models.Station.code).all()
    _write_sheet(
        wb, "tram",
        ["Mã trạm", "Tên trạm", "Đài", "Địa chỉ", "Tọa độ"],
        [[s.code, s.name, s.center, s.address, s.coordinates] for s in stations],
        [10, 26, 16, 46, 30],
    )

    employees = (
        db.query(models.Employee)
        .options(joinedload(models.Employee.station))
        .join(models.Station)
        .order_by(models.Station.code, models.Employee.name)
        .all()
    )
    _write_sheet(
        wb, "nhan vien",
        ["Họ và tên", "Mã trạm"],
        [[e.name, e.station.code] for e in employees],
        [30, 10],
    )

    categories = db.query(models.Category).order_by(models.Category.sort_order).all()
    _write_sheet(
        wb, "muc",
        ["Mục bảo dưỡng", "Mã"],
        [[c.name, c.code] for c in categories],
        [22, 8],
    )

    devices = (
        db.query(models.Device)
        .options(joinedload(models.Device.station), joinedload(models.Device.category))
        .join(models.Station)
        .join(models.Category)
        .order_by(models.Station.code, models.Category.sort_order, models.Device.name)
        .all()
    )
    device_rows = []
    for d in devices:
        station_full_name = f"{d.station.name} - {d.station.center}" if d.station.center else d.station.name
        device_rows.append([
            d.station.code, station_full_name, d.station.address, d.station.coordinates,
            d.category.name, d.name, d.record_no, d.configuration,
            "", "", "", "",
        ])
    _write_sheet(
        wb, "thiet bi",
        ["[[matram]]", "[[tentram]]", "[[diachi]]", "[[toado]]", "[[mucbaoduong]]", "[[thietbi]]",
         "[[sohs]]", "[[cauhinh]]", "[[ngaybaoduong]]", "[[thangbaoduong]]", "[[nguoithuchien]]", "[[nguoiphoihop]]"],
        device_rows,
        [10, 26, 40, 26, 16, 26, 10, 18, 14, 14, 18, 18],
    )

    return wb


def build_workbook_bytes(db: Session) -> bytes:
    wb = build_workbook(db)
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()
