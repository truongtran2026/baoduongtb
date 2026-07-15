"""Imports the legacy workbook's data sheets (tram / nhan vien / muc / thiet bi)
into the app database, and seeds the Word templates from the old Temp folder.

Re-running an import is safe: rows are upserted by their natural key
(station code, category code, employee name+station, device
station+category+name) instead of being duplicated.

For devices and employees specifically, each station (and station+category,
for devices) present in the sheet is treated as the *complete* current list
for that scope: rows that disappear from the source get deleted outright
instead of left behind - see import_devices() / import_employees(). Neither
table is referenced by history (generation runs only keep a text snapshot of
names), so this is safe. Stations and categories are never auto-deleted this
way since real history rows do reference them by id.
"""
import shutil
import uuid
from dataclasses import dataclass, field
from pathlib import Path

import openpyxl
from sqlalchemy.orm import Session

from . import models

FOLDER_AND_PREFIX_BY_CODE = {
    "VT": ("Vo tuyen", "VT-C3-M"),
    "TD": ("Truyen dan", "TD-C3-M"),
    "CM": ("Chuyen mach", "CM-C3-M"),
    "IP": ("IP bang rong", "IP-C3-M"),
}

TEMPLATE_FILE_BY_CODE = {
    "VT": "temp_vt.docx",
    "TD": "temp_td.docx",
    "CM": "temp_cm.docx",
    "IP": "temp_ip.docx",
}


@dataclass
class ImportSummary:
    stations: int = 0
    employees: int = 0
    employees_deleted: int = 0
    categories: int = 0
    devices: int = 0
    devices_deleted: int = 0
    templates: int = 0
    warnings: list[str] = field(default_factory=list)


def _cell_str(ws, row, col) -> str:
    v = ws.cell(row=row, column=col).value
    return str(v).strip() if v is not None else ""


def import_stations(db: Session, ws, summary: ImportSummary) -> dict[str, models.Station]:
    by_code: dict[str, models.Station] = {}
    row = 2
    while True:
        code = _cell_str(ws, row, 1)
        if not code:
            break
        name = _cell_str(ws, row, 2)
        center = _cell_str(ws, row, 3)
        address = _cell_str(ws, row, 4)
        coordinates = _cell_str(ws, row, 5)

        station = db.query(models.Station).filter_by(code=code).one_or_none()
        if station is None:
            station = models.Station(code=code)
            db.add(station)
        station.name = name
        station.center = center
        station.address = address
        station.coordinates = coordinates
        station.active = True
        by_code[code] = station
        summary.stations += 1
        row += 1
    db.flush()
    return by_code


def import_employees(db: Session, ws, stations_by_code: dict[str, models.Station], summary: ImportSummary):
    seen_names_by_station: dict[int, set[str]] = {}
    row = 2
    while True:
        name = _cell_str(ws, row, 1)
        station_code = _cell_str(ws, row, 2)
        if not name and not station_code:
            break
        if name:
            station = stations_by_code.get(station_code)
            if station is None:
                summary.warnings.append(f"Nhân viên '{name}': không tìm thấy trạm '{station_code}', bỏ qua")
            else:
                seen_names_by_station.setdefault(station.id, set()).add(name)
                emp = (
                    db.query(models.Employee)
                    .filter_by(name=name, station_id=station.id)
                    .one_or_none()
                )
                if emp is None:
                    emp = models.Employee(name=name, station_id=station.id)
                    db.add(emp)
                summary.employees += 1
        row += 1
    db.flush()

    # Same "sheet is the full list for what it covers" rule as devices: a
    # station whose employees appear in this sheet has its removed employees
    # deleted, not left behind.
    for station_id, names in seen_names_by_station.items():
        stale = (
            db.query(models.Employee)
            .filter(models.Employee.station_id == station_id, models.Employee.name.notin_(names))
            .all()
        )
        for e in stale:
            db.delete(e)
            summary.employees_deleted += 1
    db.flush()


def import_categories(db: Session, ws, summary: ImportSummary) -> dict[str, models.Category]:
    by_code: dict[str, models.Category] = {}
    row = 2
    order = 0
    while True:
        name = _cell_str(ws, row, 1)
        code = _cell_str(ws, row, 2)
        if not name or not code:
            break
        folder, prefix = FOLDER_AND_PREFIX_BY_CODE.get(code, (name, f"{code}-C3-M"))

        cat = db.query(models.Category).filter_by(code=code).one_or_none()
        if cat is None:
            cat = models.Category(code=code)
            db.add(cat)
        cat.name = name
        cat.output_folder_name = folder
        cat.filename_prefix = prefix
        cat.sort_order = order
        cat.active = True
        by_code[code] = cat
        by_code[name] = cat  # devices reference categories by name, not code
        summary.categories += 1
        order += 1
        row += 1
    db.flush()
    return by_code


def import_devices(db: Session, ws, stations_by_code: dict[str, models.Station],
                    categories_by_key: dict[str, models.Category], summary: ImportSummary):
    # Tracks which (station, category) combos this sheet actually covers, and
    # which device names were seen for each - used below to delete devices
    # that were removed from the source instead of leaving stale rows behind.
    seen_names: dict[tuple[int, int], set[str]] = {}
    row = 2
    while True:
        station_code = _cell_str(ws, row, 1)
        if not station_code:
            break
        category_name = _cell_str(ws, row, 5)
        device_name = _cell_str(ws, row, 6)
        record_no = _cell_str(ws, row, 7)
        configuration = _cell_str(ws, row, 8)

        station = stations_by_code.get(station_code)
        category = categories_by_key.get(category_name)
        if station is None:
            summary.warnings.append(f"Thiết bị '{device_name}': không tìm thấy trạm '{station_code}', bỏ qua")
            row += 1
            continue
        if category is None:
            # Defensive fallback: keep the device instead of silently dropping it.
            category = models.Category(
                name=category_name or "Khác",
                code=(category_name or "KH")[:10].upper(),
                output_folder_name=category_name or "Khac",
                filename_prefix=f"{(category_name or 'KH')[:4].upper()}-C3-M",
                sort_order=99,
            )
            db.add(category)
            db.flush()
            categories_by_key[category_name] = category
            summary.warnings.append(f"Mục bảo dưỡng '{category_name}' chưa khai báo, đã tự tạo mới")

        if device_name:
            seen_names.setdefault((station.id, category.id), set()).add(device_name)
            device = (
                db.query(models.Device)
                .filter_by(station_id=station.id, category_id=category.id, name=device_name)
                .one_or_none()
            )
            if device is None:
                device = models.Device(station_id=station.id, category_id=category.id, name=device_name)
                db.add(device)
            device.record_no = record_no
            device.configuration = configuration
            summary.devices += 1
        row += 1
    db.flush()

    # A (station, category) combo that appears in this sheet is treated as the
    # authoritative full list for that combo: anything still in the DB for
    # that combo but no longer present in the sheet has been removed/renamed
    # at the source, so delete it outright.
    for (station_id, category_id), names in seen_names.items():
        stale = (
            db.query(models.Device)
            .filter(
                models.Device.station_id == station_id,
                models.Device.category_id == category_id,
                models.Device.name.notin_(names),
            )
            .all()
        )
        for d in stale:
            db.delete(d)
            summary.devices_deleted += 1
    db.flush()


def import_from_workbook(db: Session, xlsx_path: Path) -> ImportSummary:
    summary = ImportSummary()
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    sheet_map = {name.lower(): name for name in wb.sheetnames}

    def get_sheet(expected: str):
        real_name = sheet_map.get(expected.lower())
        return wb[real_name] if real_name else None

    ws_tram = get_sheet("tram")
    ws_nhanvien = get_sheet("nhan vien")
    ws_muc = get_sheet("muc")
    ws_thietbi = get_sheet("thiet bi")

    if ws_tram is None or ws_muc is None or ws_thietbi is None:
        raise ValueError(
            "File Excel thiếu sheet bắt buộc (cần có 'tram', 'muc', 'thiet bi')."
        )

    stations_by_code = import_stations(db, ws_tram, summary)
    categories_by_key = import_categories(db, ws_muc, summary)
    if ws_nhanvien is not None:
        import_employees(db, ws_nhanvien, stations_by_code, summary)
    import_devices(db, ws_thietbi, stations_by_code, categories_by_key, summary)

    db.commit()
    return summary


def seed_word_templates(db: Session, templates_src_dir: Path, templates_dest_dir: Path, summary: ImportSummary):
    templates_dest_dir.mkdir(parents=True, exist_ok=True)
    categories = {c.code: c for c in db.query(models.Category).all()}

    for code, filename in TEMPLATE_FILE_BY_CODE.items():
        category = categories.get(code)
        src = templates_src_dir / filename
        if category is None or not src.exists():
            continue

        existing_active = (
            db.query(models.WordTemplate)
            .filter_by(category_id=category.id, is_active=True)
            .one_or_none()
        )
        if existing_active is not None:
            continue  # already seeded

        stored_name = f"{uuid.uuid4().hex}.docx"
        shutil.copyfile(src, templates_dest_dir / stored_name)
        db.add(
            models.WordTemplate(
                category_id=category.id,
                original_filename=filename,
                stored_filename=stored_name,
                is_active=True,
            )
        )
        summary.templates += 1
    db.commit()
